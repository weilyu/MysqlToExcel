import pymysql
import json
import openpyxl
import codecs

settings = json.loads(open('settings.json', encoding='utf-8').read())
host = settings['host']
port = settings['port']
user = settings['user']
password = settings['password']
database = settings['database']
charset = settings['charset']
connection = pymysql.connect(host=host, port=port, user=user,
                             password=password, database=database,
                             charset=charset)
tables = settings['tables']
workbook = openpyxl.Workbook()
for table in tables:
    # if table name is longer than 31 characters, use the first 31 characters
    # as worksheet name
    if (len(table) > 31):
        print(table, end=' => ')
        table = table[:31]
        print(table)
    worksheet = workbook.create_sheet(table)
    with connection.cursor() as cursor:
        get_col_names_sql = "SELECT COLUMN_NAME " + \
            "FROM INFORMATION_SCHEMA.COLUMNS " + \
            "WHERE TABLE_SCHEMA = %s AND TABLE_NAME = %s"
        count_col = cursor.execute(get_col_names_sql, (database, table))
        col_names = cursor.fetchall()
        get_data_sql = 'select * from ' + table
        cursor.execute(get_data_sql)
        data = cursor.fetchall()
        for i in range(len(col_names)):
            worksheet.cell(row=1, column=i + 1).value = col_names[i][0]
        for row in data:
            worksheet.append(row)
workbook.remove(workbook.active)
workbook.save('test.xlsx')
input('The job is done. Press enter to exist.')
